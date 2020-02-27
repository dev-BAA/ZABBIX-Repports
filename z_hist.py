# -*- coding: utf-8 -*-
from datetime import datetime
import os, sys
import time
import calendar
import datetime
import openpyxl
import logging
sys.path.append("/root")
from pyzabbix import ZabbixAPI
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl import Workbook
from settings import *

logging.basicConfig(level=logging.ERROR,
                    format='%(asctime)s %(name)s %(levelname)s %(message)s',
                    datefmt='%d-%m %H:%M:%S',
                    filename='./logs/log_hist')
logger = logging.getLogger('-')
logger.setLevel(logging.INFO)

thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="ff0000")
border = Border(top=double, left=thin, right=thin, bottom=double)
fill = PatternFill("solid", fgColor="DDDDDD")
#fill = GradientFill(stop=("000000", "FFFFFF"))
font = Font(b=True, color="000000")
al = Alignment(horizontal="center", vertical="center")
pathe = '/mnt/prin/print.xlsx'

server = 'http://127.0.0.1/zabbix'
zapi = ZabbixAPI(server)
zapi.login(login, pswd)

print(datetime.datetime.now())
datyear = datetime.datetime.today().strftime("%Y")
datmonth = datetime.datetime.today().strftime("%m")
datmoday = datetime.datetime.today().strftime("%d")
time_till = time.mktime(datetime.datetime.now().timetuple())

def histry (item_id, s_date, s_month, s_year, s_time, e_date, e_month, e_year, e_time):
    global datt, vall, history
    dat = str(s_date) + "/" + str(s_month) + "/" + str(s_year) + str(s_time)
    time_start = time.mktime(datetime.datetime.strptime(dat, "%d/%m/%Y %H:%M").timetuple())
    time_start = (str(time_start)).split('.')[0]
    logger.info("histry  -   time_start: " + str(time_start))
    dat = str(e_date) + "/" + str(e_month) + "/" + str(e_year) + str(e_time)
    time_end = time.mktime(datetime.datetime.strptime(dat, "%d/%m/%Y %H:%M").timetuple())
    time_end = (str(time_end)).split('.')[0]
    logger.info("histry  -   time_end: " + str(time_end))
    history = zapi.history.get(itemids=[item_id],
                               time_from=time_start,
                               time_till=time_end,
                               output='extend',
                               limit='1',
                               )
    if history:
        datt = format(datetime.datetime.fromtimestamp(int(history[0]['clock'])).strftime("%x %X")) #datetime
        vall = history[0]['value']
        logger.info("Дата: " + datt + "   Распечатано на дату: " + vall)
    return datt, vall, history

wb = openpyxl.load_workbook(filename = pathe)
sheetsnames = wb.get_sheet_names()
sheetsnames.reverse()
point = sheetsnames[0]
if point != datyear:
    sheet = wb.create_sheet(title=datyear)
    sheet.cell(row=1, column=1).value = 'Подразделение'
    sheet.cell(row=1, column=2).value = 'Кабинет'
    sheet.cell(row=1, column=3).value = 'ID'
    sheet.cell(row=1, column=4).value = 'Модель'
    sheet.cell(row=1, column=5).value = 'IP'
    sheet.cell(row=1, column=6).value = 'MAK'
    sheet.cell(row=1, column=7).value = 'Январь'
    sheet.cell(row=1, column=8).value = 'Февраль'
    sheet.cell(row=1, column=9).value = 'Март'
    sheet.cell(row=1, column=10).value = 'Апрель'
    sheet.cell(row=1, column=11).value = 'Май'
    sheet.cell(row=1, column=12).value = 'Июнь'
    sheet.cell(row=1, column=13).value = 'Июль'
    sheet.cell(row=1, column=14).value = 'Август'
    sheet.cell(row=1, column=15).value = 'Сентябрь'
    sheet.cell(row=1, column=16).value = 'Октябрь'
    sheet.cell(row=1, column=17).value = 'Ноябрь'
    sheet.cell(row=1, column=18).value = 'Декабрь'
    sheet.cell(row=1, column=19).value = 'Итого за год'
    sheet.cell(row=1, column=20).value = 'Всего'
    cel = 2
    y = 1
    while y < 21:
        sheet.cell(row=1, column=y).alignment = al
        sheet.cell(row=1, column=y).font = font
        sheet.cell(row=1, column=y).fill = fill
        y = y + 1.
    while cel < 1002:
        y = 1
        sheet.cell(row=cel, column=3).value = 'U-0' + '{:03}'.format(cel - 1)
        while y < 21:
            sheet.cell(row=cel, column=y).alignment = al
            y = y + 1
        cel = cel + 1
elif point == datyear and datmonth != "01":
    sheet = wb[datyear]
elif point == datyear and datmonth == "01":
    datyear_december = str(int(datyear)-1)
    sheet = wb[datyear_december]

if pathe:
    hosts = zapi.host.get(groupids=[11,12,14,19],
                       output=["hostid", "name"]
                       )
    # Получение данных по принтерам (id принтера, кол-во отпечатанных страниц, местоположение, модель, mac адрес, ip адрес) с db Zabbix
    for point in hosts:
        hostname = point['name']
        print(hostname)
        hid = point['hostid']
        item = zapi.item.get(hostids=hid,
                        output=["itemid", "name"],
                        filter={'name':'Отпечатанно страниц'}
                        )
        iten = zapi.item.get(hostids=hid,
                        output=["lastvalue", "name"],
                        filter={'name':'Местоположение'}
                        )
        location = iten[0]['lastvalue']
        iten = zapi.item.get(hostids=hid,
                        output=["lastvalue", "name"],
                        filter={'name':'Модель'}
                        )
        mdel = iten[0]['lastvalue']
        iten = zapi.item.get(hostids=hid,
                        output=["lastvalue", "name"],
                        filter={'name':'MAC адрес'}
                        )
        maca = str(iten[0]['lastvalue'])
        ips = zapi.hostinterface.get(hostids=hid,
                        output=["ip"],
                        filter={'name':'ip'}
                        )
        ip = ips[0]['ip']

        m = j = 2
        hos = sheet.cell(row=m, column=3).value

        logger.info("--- Месяц " + datmonth + " -------------------------------------------------------------------------------")
        logger.info("МФУ: " + hostname)
        logger.info("Модель: " + mdel)
        logger.info("MAC: " + maca)
        logger.info("IP: " + ip)
        logger.info("Местоположение: " + location)
        logger.info("------------------------------")

        # Получение данных по печати за прошлый месяц
        for point in item:
            item_id = point['itemid']
            if datmonth != "01":
                lastm = calendar.monthrange(int(datyear),int(datmonth)-1)[1]
                histry (item_id, 1, str(int(datmonth)-1), datyear, " 00:00", str(lastm), str(int(datmonth)-1), datyear, " 23:30")
                datt1 = datt
                vall1 = vall
                if not history:
                    vall = 0
                histry (item_id, lastm, str(int(datmonth)-1), datyear, " 18:00", str(lastm), str(int(datmonth)-1), datyear, " 23:30")
                datt31 = datt
                vall31 = vall
                if not history:
                    while not history and lastm != 1:
                        lastm-=1
                        histry (item_id, lastm, str(int(datmonth)-1), datyear, " 18:00", str(lastm), str(int(datmonth)-1), datyear, " 23:30")
                        if history:
                            datt31 = datt
                            vall31 = vall
                if lastm == 1:
                    vall = 0
    
            if datmonth == "01":
                lastm = calendar.monthrange(int(datyear)-1,12)[1]
                histry (item_id, 1, 12, str(int(datyear)-1), " 00:00", str(lastm), 12, str(int(datyear)-1), " 23:30")
                datt1 = datt
                vall1 = vall
                if not history:
                    vall = 0
                histry (item_id, lastm, 12, str(int(datyear)-1), " 18:00", str(lastm), 12, str(int(datyear)-1), " 23:30")
                datt31 = datt
                vall31 = vall
                if not history:
                    while not history and lastm != 1:
                        lastm-=1
                        histry (item_id, lastm, 12, str(int(datyear)-1), " 18:00", str(lastm), 12, str(int(datyear)-1), " 23:30")
                        if history:
                            datt31 = datt
                            vall31 = vall
                if lastm == 1:
                    vall = 0

            # Сохранение полученной информации в Excel книгу
            while hos and j != 1:
                if hos == hostname:
                    vall_r = int(vall31) - int(vall1)
                    if (datmonth != "01"):
                        sheet.cell(row=m, column=(int(datmonth) + 5)).value = vall_r
                    elif (datmonth == "01"):
                        sheet.cell(row=m, column=18).value = vall_r
                    i = 7
                    vall_sum = 0
                    while i < 18:
                        vall_s = sheet.cell(row=m, column=i).value
                        if vall_s is None:
                            vall_s = 0
                        vall_sum = int(vall_s) + int(vall_sum)
                        i = i + 1
                    sheet.cell(row=m, column=19).value = vall_sum
                    sheet.cell(row=m, column=20).value = vall31
                    if (location != "0" and location != None and location != 0):
                        sheet.cell(row=m, column=2).value = location
                    if (mdel != "0" and mdel != None and mdel != 0):
                        sheet.cell(row=m, column=4).value = mdel
                    HTTP = "http://" + ip
                    sheet.cell(row=m, column=5).value = HTTP
                    if (maca != "0" and maca != None and maca != 0):
                        sheet.cell(row=m, column=6).value = maca
                    j = 1
                    logger.info("------------------------------")
                    logger.info("Распечатано за месяц: " + str(vall_r))
                    logger.info("Распечатано за год: " + str(vall_sum))
                m = m + 1
                hos = sheet.cell(row=m, column=3).value
            vall = 0
            datt = 0
            vall_end = 0
            datt_end = 0

sheet.column_dimensions['A'].width = 63
sheet.column_dimensions['C'].width = 8
sheet.column_dimensions['D'].width = 20
sheet.column_dimensions['E'].width = 18
sheet.column_dimensions['F'].width = 16
sheet.column_dimensions['S'].width = 15
wb.save(pathe)